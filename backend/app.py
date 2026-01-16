# app.py
# SISTEMA COMPLETO ENTERPRISE
# Inclui: JWT, Histórico, Exportação (PDF/Excel), WhatsApp e Google Sheets

from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from jose import jwt, JWTError
from datetime import datetime, timedelta
from pathlib import Path
import shutil
import csv
import json
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
import uuid
from typing import Optional, Dict, Any, List
import os
import requests
import io
from openpyxl import load_workbook, Workbook

# Serviço de cache SQLite (opcional)
try:
    from cache_service import cache_service
except ImportError:
    cache_service = None
    print("⚠️  cache_service não encontrado, continuando sem cache...")

# =========================
# CONFIG
# =========================

# Validação de SECRET_KEY em produção (com fallback para desenvolvimento)
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    SECRET_KEY = "dev-secret-key-CHANGE-IN-PRODUCTION-" + str(uuid.uuid4())
    print("⚠️  AVISO: Usando SECRET_KEY temporária! Defina SECRET_KEY em produção.")

ALGORITHM = "HS256"
TOKEN_EXPIRE_MIN = 60

app = FastAPI(
    title="Chat IA Corporativo",
    description="API Enterprise para Análise de Relatórios com IA",
    version="1.0.0"
)

security = HTTPBearer()

# CORS configurável via variável de ambiente
allowed_origins = os.getenv("ALLOWED_ORIGINS", "*").split(",")
if allowed_origins == ["*"]:
    print("⚠️  AVISO: CORS configurado para permitir todas as origens (desenvolvimento)")
else:
    print(f"✅ CORS configurado para: {', '.join(allowed_origins)}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path("data")
BASE_DIR.mkdir(exist_ok=True)
LOGS_FILE = BASE_DIR / "logs.csv"
UPLOADS_DIR = BASE_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)
EXPORTS_DIR = BASE_DIR / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)

# Inicializar arquivo de logs
if not LOGS_FILE.exists():
    with open(LOGS_FILE, "w", encoding="utf-8") as f:
        f.write("timestamp,usuario,tipo,codvd,vendedor,registros\n")

# Banco de dados de usuários (em produção, usar BD real)
USERS_DB = {
    "admin@teste.com": {"password": "123456", "role": "admin", "name": "Admin Teste"},
    "teste@teste.com": {"password": "123456", "role": "user", "name": "Usuário Teste"},
    "nathiely@empresa.com": {"password": "Nathiely@2025", "role": "admin", "name": "Nathiely"},
    "roberto.felix@empresa.com": {"password": "Roberto@2025", "role": "admin", "name": "Roberto Felix"},
}

# =========================
# JWT
# =========================

def criar_token(email: str, role: str = "user"):
    payload = {
        "sub": email,
        "role": role,
        "exp": datetime.utcnow() + timedelta(minutes=TOKEN_EXPIRE_MIN)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def get_user(token: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(token.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        return {"email": payload["sub"], "role": payload.get("role", "user")}
    except JWTError:
        raise HTTPException(403, "Token inválido ou expirado")

# =========================
# AUTH
# =========================

@app.post("/api/auth/login")
def login(credentials: Dict[str, Any] = Body(...)):
    email = credentials.get("email")
    password = credentials.get("password")
    
    if not email or not password:
        raise HTTPException(400, "Email e senha são obrigatórios")
    
    user = USERS_DB.get(email)
    if not user or user["password"] != password:
        raise HTTPException(401, "Credenciais inválidas")
    
    token = criar_token(email, user["role"])
    
    return {
        "token": token,
        "user": {
            "email": email,
            "name": user["name"],
            "role": user["role"]
        }
    }


@app.post("/api/auth/register")
def register(credentials: Dict[str, Any] = Body(...)):
    email = credentials.get("email")
    password = credentials.get("password")
    name = credentials.get("name", "Novo Usuário")
    
    if not email or not password:
        raise HTTPException(400, "Email e senha são obrigatórios")
    
    if email in USERS_DB:
        raise HTTPException(400, "Usuário já existe")
    
    USERS_DB[email] = {
        "password": password,
        "role": "user",
        "name": name
    }
    
    token = criar_token(email, "user")
    
    return {
        "token": token,
        "user": {
            "email": email,
            "name": name,
            "role": "user"
        }
    }


@app.get("/api/auth/me")
def me(user_data = Depends(get_user)):
    email = user_data["email"]
    user = USERS_DB.get(email)
    
    return {
        "email": email,
        "name": user["name"],
        "role": user["role"]
    }

# =========================
# HISTÓRICO
# =========================

def salvar_log(usuario, tipo, codvd, vendedor, registros):
    linha = {
        "timestamp": datetime.now().isoformat(),
        "usuario": usuario,
        "tipo": tipo,
        "codvd": codvd,
        "vendedor": vendedor,
        "registros": registros
    }
    
    with open(LOGS_FILE, "a", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["timestamp", "usuario", "tipo", "codvd", "vendedor", "registros"])
        writer.writerow(linha)


@app.get("/api/historico")
def obter_historico(user_data = Depends(get_user)):
    if not LOGS_FILE.exists():
        return {"historico": []}
    
    # Ler CSV com csv module nativo
    historico = []
    with open(LOGS_FILE, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Filtrar por usuário se não for admin
            if user_data["role"] == "admin" or row.get("usuario") == user_data["email"]:
                historico.append(row)
    
    # Últimos 100 registros
    historico = historico[-100:]
    
    return {"historico": historico}

# =========================
# UPLOAD DE PLANILHAS
# =========================

@app.post("/api/upload/excel")
def upload_excel(file: UploadFile = File(...), user_data = Depends(get_user)):
    # Apenas admins podem fazer upload
    if user_data["role"] != "admin":
        raise HTTPException(403, "Apenas administradores podem fazer upload")
    
    # Validar extensão
    if not file.filename.endswith(('.xlsx', '.xls', '.xlsm', '.csv')):
        raise HTTPException(400, "Arquivo deve ser Excel ou CSV")
    
    # Salvar arquivo
    file_id = str(uuid.uuid4())
    ext = Path(file.filename).suffix
    file_path = UPLOADS_DIR / f"{file_id}{ext}"
    
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    
    # Tentar ler para validar
    try:
        if ext == '.csv':
            # Ler CSV com csv module
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                columns = reader.fieldnames if reader.fieldnames else []
        else:
            # Ler Excel com openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active
            rows = list(ws.values)
            columns = rows[0] if rows else []
            rows = len(rows) - 1 if rows else 0
        
        return {
            "file_id": file_id,
            "filename": file.filename,
            "rows": len(rows) if isinstance(rows, list) else rows,
            "columns": list(columns) if columns else [],
            "path": str(file_path)
        }
    except Exception as e:
        if file_path.exists():
            file_path.unlink()  # Deletar arquivo inválido
        raise HTTPException(400, f"Erro ao ler arquivo: {str(e)}")

# =========================
# EXPORTAÇÃO
# =========================

def exportar_excel(data: List[Dict], filename: str):
    """Exportar dados para Excel usando openpyxl diretamente"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    path = EXPORTS_DIR / f"{filename}_{uuid.uuid4().hex[:8]}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatório'
    
    if not data:
        wb.save(path)
        return path
    
    # Cabeçalho
    headers = list(data[0].keys())
    ws.append(headers)
    
    # Estilizar cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Dados
    for row in data:
        ws.append([row.get(h, '') for h in headers])
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(path)
    return path


def exportar_pdf(data: List[Dict], titulo: str, filename: str):
    """Exportar dados para PDF"""
    path = EXPORTS_DIR / f"{filename}_{uuid.uuid4().hex[:8]}.pdf"
    
    doc = SimpleDocTemplate(str(path), pagesize=A4)
    styles = getSampleStyleSheet()
    elementos = []
    
    # Título
    elementos.append(Paragraph(titulo, styles['Title']))
    elementos.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elementos.append(Paragraph(f"Total de registros: {len(data)}", styles['Normal']))
    elementos.append(Paragraph("<br/><br/>", styles['Normal']))
    
    if not data:
        doc.build(elementos)
        return path
    
    # Tabela - limitar a 50 registros e 6 colunas para PDF
    max_cols = 6
    max_rows = 50
    
    # Limitar colunas
    headers_limited = headers[:max_cols]
    data_limited = data[:max_rows]
    
    table_data = [headers_limited]
    for row in data_limited:
        table_data.append([str(row.get(h, ''))[:30] for h in headers_limited])  # Limitar texto
    
    # Criar tabela
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elementos.append(table)
        
        if len(df) > 50:
            elementos.append(Paragraph(f"<br/>Mostrando primeiros 50 de {len(df)} registros", styles['Normal']))
    
    doc.build(elementos)
    return path

# =========================
# RELATÓRIOS
# =========================

@app.post("/api/relatorios/gerar")
def gerar_relatorio(payload: Dict[str, Any] = Body(...), user_data = Depends(get_user)):
    tipo = payload.get("tipo")
    codvd = payload.get("codvd")
    vendedor = payload.get("vendedor", "")
    exportar = payload.get("exportar", "json")  # json, excel, pdf
    
    if not tipo or not codvd:
        raise HTTPException(400, "Tipo e CODVD são obrigatórios")
    
    # Buscar arquivo de dados correspondente
    # Por padrão, busca arquivos .xlsx com nome do tipo
    arquivo_mapeamento = {
        "nao_cobertos_clientes": "nao_cobertos.xlsx",
        "nao_cobertos_fornecedor": "nao_cobertos.xlsx",
        "msl_mini": "msl.xlsx",
        "msl_super": "msl.xlsx",
        "msl_otg": "msl.xlsx",
        "msl_danone": "msl.xlsx",
        "exp": "msl.xlsx",
        "novos_clientes": "novos_clientes.xlsx",
        "queijo_reino": "queijo_reino.xlsx",
    }
    
    arquivo_nome = arquivo_mapeamento.get(tipo)
    if not arquivo_nome:
        raise HTTPException(400, f"Tipo de relatório desconhecido: {tipo}")
    
    arquivo_path = UPLOADS_DIR / arquivo_nome
    
    if not arquivo_path.exists():
        raise HTTPException(404, f"Arquivo de dados não encontrado: {arquivo_nome}. Faça upload primeiro.")
    
    # Ler dados usando openpyxl
    try:
        from openpyxl import load_workbook
        wb = load_workbook(arquivo_path)
        ws = wb.active
        
        # Converter para lista de dicts
        headers = [cell.value for cell in ws[1]]
        data_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data_list.append(dict(zip(headers, row)))
    except Exception as e:
        raise HTTPException(500, f"Erro ao ler arquivo: {str(e)}")
    
    # Aplicar filtros
    filtered_data = []
    for row in data_list:
        status = str(row.get("STATUS", "")).upper().strip()
        codvd_val = str(row.get("CODVD", "")).strip()
        vendedor_val = str(row.get("VENDEDOR", "")).upper()
        
        # Filtro de status por tipo
        if tipo.startswith("nao_cobertos"):
            if status != "FALTA":
                continue
        elif tipo.startswith("msl") or tipo == "exp":
            if status not in ["OK", "FALTA"]:
                continue
        
        # Filtrar por CODVD
        if codvd_val != str(codvd).strip():
            continue
        
        # Filtrar por vendedor se fornecido
        if vendedor and vendedor.upper() not in vendedor_val:
            continue
        
        filtered_data.append(row)
    
    # Remover duplicatas (simples - converte para tuplas e usa set)
    seen = set()
    unique_data = []
    for row in filtered_data:
        row_tuple = tuple(sorted(row.items()))
        if row_tuple not in seen:
            seen.add(row_tuple)
            unique_data.append(row)
    
    # Salvar log
    salvar_log(user_data["email"], tipo, codvd, vendedor, len(unique_data))
    
    # Exportar
    if exportar == "excel":
        path = exportar_excel(unique_data, tipo)
        return FileResponse(path, filename=f"{tipo}.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    elif exportar == "pdf":
        path = exportar_pdf(unique_data, f"Relatório {tipo.upper()}", tipo)
        return FileResponse(path, filename=f"{tipo}.pdf", media_type="application/pdf")
    
    else:
        # JSON
        return {
            "tipo": tipo,
            "codvd": codvd,
            "vendedor": vendedor,
            "total_registros": len(unique_data),
            "dados": unique_data
        }

# =========================
# WHATSAPP
# =========================

@app.post("/api/whatsapp/enviar")
def enviar_whatsapp(payload: Dict[str, Any] = Body(...), user_data = Depends(get_user)):
    telefone = payload.get("telefone")
    mensagem = payload.get("mensagem")
    
    if not telefone or not mensagem:
        raise HTTPException(400, "Telefone e mensagem são obrigatórios")
    
    # Aqui você integraria com Twilio, Z-API, Meta WhatsApp Business API, etc.
    # Por enquanto, apenas simula o envio
    
    print(f"📱 WhatsApp para {telefone}: {mensagem}")
    
    return {
        "status": "enviado",
        "telefone": telefone,
        "timestamp": datetime.now().isoformat()
    }

# =========================
# HEALTH CHECK
# =========================

@app.get("/")
def root():
    return {
        "app": "Chat IA Corporativo",
        "version": "1.0.0",
        "status": "online",
        "endpoints": {
            "auth": "/api/auth/login",
            "relatorios": "/api/relatorios/gerar",
            "upload": "/api/upload/excel",
            "historico": "/api/historico",
            "whatsapp": "/api/whatsapp/enviar"
        }
    }


# =========================
# GOOGLE SHEETS INTEGRATION
# =========================

# Configuração dos relatórios do Google Sheets
REPORTS_CONFIG = [
    {
        "id": "leads",
        "label": "Novos Clientes",
        "keywords": ["novos", "cidade", "leads"],
        "type": "city_leads",
        "url": "https://docs.google.com/spreadsheets/d/e/2PACX-1vR9lG9sbtgRqV0PLkyjT8R9znpC9ECGurgfelIhn_q5BwgThg6SpdfE2R30obAAaawk0FIGLlBowjt_/pub?gid=0&single=true&output=csv"
    },
    {
        "id": "queijo",
        "label": "Queijo do Reino",
        "keywords": ["queijo", "reino"],
        "type": "client_code_details",
        "url": "https://docs.google.com/spreadsheets/d/e/2PACX-1vR9lG9sbtgRqV0PLkyjT8R9znpC9ECGurgfelIhn_q5BwgThg6SpdfE2R30obAAaawk0FIGLlBowjt_/pub?gid=1824827366&single=true&output=csv"
    },
    {
        "id": "nao_cobertos_fornecedor",
        "label": "Não Cobertos (Fornecedor)",
        "keywords": ["não", "cobertos", "fornecedor"],
        "type": "supplier_coverage",
        "url": "https://docs.google.com/spreadsheets/d/e/2PACX-1vR9lG9sbtgRqV0PLkyjT8R9znpC9ECGurgfelIhn_q5BwgThg6SpdfE2R30obAAaawk0FIGLlBowjt_/pub?gid=1981950621&single=true&output=csv"
    }
]

# Cache em memória para os dados das planilhas
report_data_cache: Dict[str, List[Dict]] = {}
report_validation_status: Dict[str, Dict] = {}
is_loading_sheets = False
last_update_time = None

# Schemas esperados para validação
REPORT_SCHEMAS = {
    "leads": {
        "version": 1,
        "columns": ["Cidade", "Novos Clientes", "Data"]
    },
    "queijo": {
        "version": 1,
        "columns": ["Código Cliente", "Nome", "Detalhes"]
    },
    "nao_cobertos_fornecedor": {
        "version": 1,
        "columns": ["Fornecedor", "Produto", "Status", "Observações"]
    }
}


def validate_report_schema(report_id: str, data: List[Dict]) -> Dict:
    """Valida se os dados correspondem ao schema esperado"""
    if not data:
        return {"ok": False, "error": "Dados vazios"}
    
    schema = REPORT_SCHEMAS.get(report_id)
    if not schema:
        return {"ok": True, "warning": "Schema não definido"}
    
    headers = list(data[0].keys())
    expected_columns = schema["columns"]
    missing_columns = [col for col in expected_columns if col not in headers]
    extra_columns = [col for col in headers if col not in expected_columns]
    
    return {
        "ok": len(missing_columns) == 0,
        "version": schema["version"],
        "missing_columns": missing_columns,
        "extra_columns": extra_columns,
        "expected_columns": expected_columns,
        "actual_columns": headers
    }


def parse_csv_text(text: str) -> List[Dict[str, str]]:
    """Parser de CSV robusto"""
    lines = [line for line in text.split('\n') if line.strip()]
    if not lines:
        return []
    
    reader = csv.DictReader(io.StringIO('\n'.join(lines)))
    return [row for row in reader]


async def carregar_dados_sheets(force_refresh: bool = False):
    """Carrega dados de todas as planilhas configuradas
    
    Args:
        force_refresh: Se True, ignora cache e busca do Google Sheets
    """
    global is_loading_sheets, last_update_time
    is_loading_sheets = True
    print("📥 Carregando planilhas do Google Sheets...")
    
    # Se não forçar, tenta usar cache (24h)
    if not force_refresh:
        all_fresh = True
        for config in REPORTS_CONFIG:
            if not cache_service.is_cache_fresh(config["id"], max_age_hours=24):
                all_fresh = False
                break
        
        if all_fresh:
            print("✅ Usando dados do cache (atualizados nas últimas 24h)")
            for config in REPORTS_CONFIG:
                cached = cache_service.get_report_cache(config["id"])
                if cached:
                    report_data_cache[config["id"]] = cached["data"]
                    report_validation_status[config["id"]] = cached.get("validation_status", {"ok": True})
                    print(f"  📋 {config['label']}: {cached['row_count']} linhas (cache)")
            
            is_loading_sheets = False
            last_update_time = datetime.now().isoformat()
            print("🟢 Carga concluída via cache")
            return report_data_cache
    
    for config in REPORTS_CONFIG:
        try:
            response = requests.get(config["url"], timeout=10)
            response.raise_for_status()
            
            text = response.text
            data = parse_csv_text(text)
            
            # Validar schema
            validation = validate_report_schema(config["id"], data)
            report_validation_status[config["id"]] = validation
            
            if validation["ok"]:
                report_data_cache[config["id"]] = data
                print(f"✅ {config['label']} carregado ({len(data)} linhas) - Schema v{validation['version']} OK")
            else:
                report_data_cache[config["id"]] = data  # Carrega mesmo com erro
                print(f"⚠️ {config['label']} carregado ({len(data)} linhas) - Schema inválido")
                print(f"   Colunas faltando: {validation.get('missing_columns', [])}")
                if validation.get('extra_columns'):
                    print(f"   Colunas extras: {validation['extra_columns']}")
            
            # Salvar no cache SQLite
            cache_service.save_report_cache(
                report_id=config["id"],
                label=config["label"],
                data=data,
                validation_status=validation
            )
            
        except Exception as e:
            print(f"❌ Falha em {config['label']}: {str(e)}")
            
            # Tentar buscar do cache como fallback
            cached = cache_service.get_report_cache(config["id"])
            if cached:
                print(f"   📦 Usando versão em cache ({cached['row_count']} linhas)")
                report_data_cache[config["id"]] = cached["data"]
                report_validation_status[config["id"]] = cached.get("validation_status", {"ok": False})
            else:
                report_data_cache[config["id"]] = []
    
    is_loading_sheets = False
    last_update_time = datetime.now().isoformat()
    print("🟢 Carga finalizada")
    return report_data_cache


@app.on_event("startup")
async def startup_event():
    """Carrega dados das planilhas ao iniciar o servidor"""
    await carregar_dados_sheets()


@app.get("/api/sheets/reload")
def reload_sheets(force: bool = True, user: dict = Depends(get_user)):
    """Recarrega dados das planilhas do Google Sheets
    
    Args:
        force: Se True (padrão), ignora cache e busca do Google Sheets
    """
    try:
        import asyncio
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        result = loop.run_until_complete(carregar_dados_sheets(force_refresh=force))
        
        summary = {
            config["id"]: {
                "label": config["label"],
                "rows": len(report_data_cache.get(config["id"], []))
            }
            for config in REPORTS_CONFIG
        }
        
        return {
            "status": "success",
            "message": "Dados recarregados com sucesso",
            "timestamp": datetime.now().isoformat(),
            "data": summary
        }
    except Exception as e:
        raise HTTPException(500, f"Erro ao recarregar planilhas: {str(e)}")


@app.get("/api/sheets/{report_id}")
def get_sheet_data(report_id: str, user: dict = Depends(get_user)):
    """Retorna dados de uma planilha específica"""
    if report_id not in report_data_cache:
        raise HTTPException(404, f"Relatório '{report_id}' não encontrado")
    
    data = report_data_cache[report_id]
    validation = report_validation_status.get(report_id, {"ok": True})
    
    return {
        "id": report_id,
        "data": data,
        "count": len(data),
        "validation": validation,
        "timestamp": last_update_time
    }
    config = next((c for c in REPORTS_CONFIG if c["id"] == report_id), None)
    
    return {
        "id": report_id,
        "label": config["label"] if config else report_id,
        "rows": len(data),
        "data": data,
        "timestamp": datetime.now().isoformat()
    }


@app.get("/api/status")
def get_status():
    """Retorna status do carregamento das planilhas"""
    return {
        "loading": is_loading_sheets,
        "lastUpdate": last_update_time,
        "reports": list(report_data_cache.keys())
    }


@app.get("/api/sheets")
def list_sheets(user: dict = Depends(get_user)):
    """Lista todas as planilhas disponíveis com status de validação"""
    sheets = []
    for config in REPORTS_CONFIG:
        data = report_data_cache.get(config["id"], [])
        validation = report_validation_status.get(config["id"], {"ok": True})
        
        sheets.append({
            "id": config["id"],
            "label": config["label"],
            "keywords": config["keywords"],
            "type": config["type"],
            "rows": len(data),
            "has_data": len(data) > 0,
            "validation": {
                "ok": validation.get("ok", True),
                "version": validation.get("version"),
                "issues": validation.get("missing_columns", [])
            }
        })
    
    return {
        "sheets": sheets,
        "total": len(sheets),
        "timestamp": datetime.now().isoformat()
    }


@app.get("/api/health")
def api_health():
    """Endpoint de saúde com informações detalhadas dos relatórios"""
    reports_status = {}
    
    for config in REPORTS_CONFIG:
        data = report_data_cache.get(config["id"], [])
        reports_status[config["id"]] = {
            "ok": len(data) > 0,
            "rows": len(data),
            "lastUpdate": last_update_time,
            "label": config["label"]
        }
    
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "loading": is_loading_sheets,
        "reports": reports_status
    }


@app.get("/api/cache/info")
def cache_info(user: dict = Depends(get_user)):
    """Retorna informações sobre os dados em cache"""
    try:
        cached_reports = cache_service.list_cached_reports()
        history = cache_service.get_update_history(limit=20)
        
        return {
            "status": "success",
            "timestamp": datetime.now().isoformat(),
            "cached_reports": cached_reports,
            "total_cached": len(cached_reports),
            "recent_updates": history,
            "database_path": str(cache_service.db_path)
        }
    except Exception as e:
        raise HTTPException(500, f"Erro ao buscar informações do cache: {str(e)}")


@app.post("/api/cache/clear")
def clear_cache(days_old: int = 30, user: dict = Depends(get_user)):
    """Remove caches mais antigos que X dias
    
    Args:
        days_old: Idade em dias para considerar cache obsoleto (padrão: 30)
    """
    try:
        cache_service.clear_old_cache(days_old)
        return {
            "status": "success",
            "message": f"Caches com mais de {days_old} dias foram removidos",
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        raise HTTPException(500, f"Erro ao limpar cache: {str(e)}")


@app.get("/health")
def health():
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}



if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
